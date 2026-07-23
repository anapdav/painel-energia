# Coletor EPE — consumo mensal de energia elétrica por classe (Brasil, desde 2004)
# XLSX único com blocos verticais por ano (linha do ano -> linha dos meses ->
# linha "TOTAL BRASIL"). Anos com "*" = dados preliminares (rotulado na fonte).
import io
import urllib.request

import openpyxl

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
URL = ("https://www.epe.gov.br/sites-pt/publicacoes-dados-abertos/publicacoes/"
       "Documents/CONSUMO%20MENSAL%20DE%20ENERGIA%20EL%C3%89TRICA%20POR%20CLASSE.xlsx")
ABAS = {"TOTAL": "total", "RESIDENCIAL": "residencial", "INDUSTRIAL": "industrial",
        "COMERCIAL": "comercial", "CATIVO": "cativo", "LIVRE": "livre"}


def _pares_da_aba(ws):
    """Percorre blocos anuais; devolve [(YYYY-MM-01, valor MWh do TOTAL BRASIL)]."""
    linhas = list(ws.iter_rows(values_only=True))
    pares = []
    for i, row in enumerate(linhas):
        ano_bruto = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        ano = ano_bruto.replace("*", "").split(".")[0]
        if not (ano.isdigit() and len(ano) == 4):
            continue
        # linha seguinte deve ser JAN..DEZ; procura TOTAL BRASIL nas 4 seguintes
        for j in range(i + 1, min(i + 6, len(linhas))):
            r = linhas[j]
            # linha-total: "TOTAL BRASIL" na aba TOTAL, "TOTAL RESIDENCIAL" etc nas demais
            if r and r[0] is not None and str(r[0]).strip().upper().startswith("TOTAL"):
                for mes in range(1, 13):
                    v = r[mes] if len(r) > mes else None
                    if isinstance(v, (int, float)) and v > 0:
                        pares.append((f"{ano}-{mes:02d}-01", float(v)))
                break
    return sorted(pares)


def coleta(con, registra_serie, grava_dados, log=print):
    raw = urllib.request.urlopen(urllib.request.Request(URL, headers=UA), timeout=300).read()
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    for aba, sufixo in ABAS.items():
        if aba not in wb.sheetnames:
            log(f"  [AVISO] epe: aba {aba} não encontrada")
            continue
        pares = _pares_da_aba(wb[aba])
        sid = f"epe_consumo_{sufixo}"
        registra_serie(con, sid, "EPE", "BR", "eletricidade",
                       f"Consumo de energia elétrica na rede — {aba.title()}, Brasil "
                       "(meses recentes preliminares)", "MWh/mês", "mensal", URL)
        grava_dados(con, sid, pares)
        log(f"  {sid}: {len(pares)} meses | ultimo {pares[-1][0] if pares else '-'}")
