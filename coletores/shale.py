# Coletor Shale EUA — Drilling Productivity Report (DPR) + DUCs da EIA
# DPR (mensal, 7 bacias, desde 2007): rigs, produtividade do poço novo por
# sonda, produção total de óleo (bbl/d) e gás (Mcf/d) por região.
# DUC (mensal, desde 2013): poços perfurados, completados e DUCs (perfurados
# não completados) por região. "Novas licenças" (permits) NÃO são dado EIA
# (agências estaduais); o proxy analítico aqui é perfurados/completados.
# + produção mensal de crude por estado (TX/NM/ND) via API.
import io
import urllib.request

import openpyxl

from coletores.eia import _fetch

UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
URL_DPR = "https://www.eia.gov/petroleum/drilling/xls/dpr-data.xlsx"
URL_DUC = "https://www.eia.gov/petroleum/drilling/xls/duc-data.xlsx"

REGIOES = {"Anadarko Region": "anadarko", "Appalachia Region": "appalachia",
           "Bakken Region": "bakken", "Eagle Ford Region": "eagleford",
           "Haynesville Region": "haynesville", "Niobrara Region": "niobrara",
           "Permian Region": "permian"}
NOMES = {"anadarko": "Anadarko", "appalachia": "Appalachia", "bakken": "Bakken",
         "eagleford": "Eagle Ford", "haynesville": "Haynesville",
         "niobrara": "Niobrara", "permian": "Permian"}

ESTADOS = [("MCRFPTX2", "tx", "Texas"), ("MCRFPNM2", "nm", "Novo México"),
           ("MCRFPND2", "nd", "Dakota do Norte")]


def _xlsx(url):
    raw = urllib.request.urlopen(urllib.request.Request(url, headers=UA),
                                 timeout=300).read()
    return openpyxl.load_workbook(io.BytesIO(raw), read_only=True)


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def coleta(con, registra_serie, grava_dados, log=print):
    # --- DPR: produção e produtividade por bacia ---
    try:
        wb = _xlsx(URL_DPR)
        for aba, slug in REGIOES.items():
            rigs, novo, oleo, gas = [], [], [], []
            for row in wb[aba].iter_rows(min_row=3, values_only=True):
                if not row or row[0] is None:
                    continue
                d = str(row[0])[:10]
                if _f(row[1]) is not None:
                    rigs.append((d, _f(row[1])))
                if _f(row[2]) is not None:
                    novo.append((d, _f(row[2])))
                if _f(row[4]) is not None:
                    oleo.append((d, _f(row[4]) / 1000.0))      # bbl/d -> mil b/d
                if _f(row[7]) is not None:
                    gas.append((d, _f(row[7]) / 1e6))          # Mcf/d -> Bcf/d
            nome = NOMES[slug]
            FIM = " (DPR — relatório DESCONTINUADO pela EIA; última edição jun/2024)"
            for sid, pares, desc, unid in [
                (f"shale_{slug}_rigs", rigs, f"Rigs ativos — {nome}{FIM}", "rigs"),
                (f"shale_{slug}_novo_poco", novo,
                 f"Produção de óleo do poço novo por sonda — {nome}{FIM}", "b/d por rig"),
                (f"shale_{slug}_oleo", oleo, f"Produção de óleo — {nome}{FIM}", "mil b/d"),
                (f"shale_{slug}_gas", gas,
                 f"Produção de gás — {nome}, convertido de Mcf/d{FIM}", "Bcf/d"),
            ]:
                registra_serie(con, sid, "EIA", "US", "shale", desc, unid,
                               "mensal", URL_DPR)
                grava_dados(con, sid, pares)
        log(f"  shale_dpr: {len(REGIOES)} bacias x 4 series | "
            f"ultimo {oleo[-1][0] if oleo else '-'}")
    except Exception as e:
        log(f"  [ERRO] shale_dpr: {type(e).__name__}: {e}")

    # --- DUC: perfurados, completados e DUCs por bacia ---
    try:
        ws = _xlsx(URL_DUC)["Data"]
        linhas = list(ws.iter_rows(values_only=True))
        cab_reg, cab_sub = linhas[2], linhas[3]
        blocos = []                       # (col_inicial, slug)
        for j, nome in enumerate(cab_reg):
            slug = REGIOES.get(f"{str(nome).strip()} Region") if nome else None
            if slug and str(cab_sub[j]).strip() == "Drilled":
                blocos.append((j, slug))
        series = {}
        for row in linhas[4:]:
            if not row or row[0] is None:
                continue
            d = str(row[0])[:10]
            for j, slug in blocos:
                for off, tipo in [(0, "perfurados"), (1, "completados"), (2, "duc")]:
                    v = _f(row[j + off])
                    if v is not None:
                        series.setdefault(f"shale_{slug}_{tipo}", []).append((d, v))
        DESC = {"perfurados": "Poços perfurados", "completados": "Poços completados",
                "duc": "DUCs (perfurados não completados)"}
        for sid, pares in series.items():
            slug, tipo = sid.split("_")[1], sid.split("_")[2]
            registra_serie(con, sid, "EIA", "US", "shale",
                           f"{DESC[tipo]} — {NOMES[slug]} (relatório DUC "
                           "DESCONTINUADO; última edição abr/2024)",
                           "poços", "mensal", URL_DUC)
            grava_dados(con, sid, pares)
        log(f"  shale_duc: {len(series)} series")
    except Exception as e:
        log(f"  [ERRO] shale_duc: {type(e).__name__}: {e}")

    # --- Produção mensal de crude por estado ---
    for serie_eia, slug, nome in ESTADOS:
        try:
            linhas = _fetch("petroleum/crd/crpdn", serie_eia, "monthly")
            pares = [(r["period"] + "-01", float(r["value"])) for r in linhas
                     if r.get("value") not in (None, "")]
            sid = f"shale_estado_{slug}"
            registra_serie(con, sid, "EIA", "US", "crude",
                           f"Produção de crude — {nome} (mensal)", "mil b/d",
                           "mensal", f"api.eia.gov petroleum/crd/crpdn [{serie_eia}]")
            grava_dados(con, sid, pares)
            log(f"  shale_estado_{slug}: {len(pares)} meses")
        except Exception as e:
            log(f"  [ERRO] shale_estado_{slug}: {type(e).__name__}: {e}")
